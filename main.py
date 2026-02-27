import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

# Define the connection string
conn_str = (
    'mssql+pyodbc://@SERVER,PORT/DATABASE?driver=DRIVER_NAME&trusted_connection=yes'
)

try:
    # Create the SQLAlchemy engine
    engine = create_engine(conn_str)

    # Define the query
    query = """
            SELECT name,
                type_desc,
                is_disabled,
                default_database_name
            FROM sys.server_principals
            WHERE type IN ('S', 'U', 'G');   -- SQL, Windows users, Windows groups
    """

    # Execute the query and fetch the results into a DataFrame
    df = pd.read_sql(query, engine)

    # Export the DataFrame to an Excel file
    file_path = 'UserQueryResults.xlsx'
    df.to_excel(file_path, index=False)

    # Load the workbook and select the active worksheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Define the border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Apply the border to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(file_path)

    print("Data exported successfully to UserQueryResults.xlsx with borders and adjusted column widths")

except Exception as e:
    print("Error in connection or execution:", e)